from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string



def GetMailList(FilePath,Column,RowString,ScorriNumeri):
    
    # TODO: Inserire supporto per file .csv

    try:
        WBook = load_workbook(FilePath)
    except:
        return -2

    WSheet = WBook.active

    ListaMails = []

    Row = int(RowString)

    if ScorriNumeri:
        while WSheet[ Column + str(Row) ].value != None:
            ListaMails.append(WSheet[ Column + str(Row) ].value)
            Row += 1
    else:

        column_number = column_index_from_string(Column)

        while WSheet.cell(row=Row, column=column_number).value != None:
            ListaMails.append( WSheet.cell(row=Row, column=column_number).value)
            column_number += 1


    if len(ListaMails) == 0:
        return -1
    else:
        return ListaMails